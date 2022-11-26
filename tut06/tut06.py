'''
from datetime import datetime
start_time = datetime.now()

import pandas as pd
import os
dir_name = os.path.dirname(__file__)
#def attendance_report():

###Code
df1=pd.read_excel('')
df2=pd.read_excel('')
writer=pd.ExcelWriter()

s = '28-07-2022  15:30:00'
sq = datetime.strptime(s, "%d-%m-%Y %H:%M:%S")

print(sq.time())


from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

#attendance_report()

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
'''
'''
import os
import openpyxl
from platform import python_version
from datetime import datetime
start_time = datetime.now()

ver = python_version()

os.system("cls")

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

inputAttendance = dir_name+"/input_attendance.csv"
inputRegisteredFile = dir_name+"/input_registered_students.csv"

roll_to_name = {}
roll_attendance = {}
dates = []


def consolidate_attendance_func():
    try:
        outputFileName = dir_name+"output/attendance_report_consolidated.xlsx"
        outputFile = openpyxl.Workbook()
        outputSheet = outputFile.active

        outputSheet.cell(row=1, column=1).value = "Roll"
        outputSheet.cell(row=1, column=2).value = "Name"

        last = -1
        for i, date in enumerate(dates):
            outputSheet.cell(row=1, column=3+i).value = date
            last = i+3

        last += 1
        list = ["Actual Lecture Taken", "Total Real", "% Attendance"]

        for i, title in enumerate(list):
            outputSheet.cell(row=1, column=last+i).value = title

        for i, rollNum in enumerate(roll_to_name.keys()):
            outputSheet.cell(row=i+2, column=1).value = rollNum
            outputSheet.cell(row=i+2, column=2).value = roll_to_name[rollNum]

            present = 0
            for j, date in enumerate(dates):
                if date not in roll_attendance[rollNum]:
                    outputSheet.cell(row=i+2, column=j+3).value = "A"
                else:
                    list = roll_attendance[rollNum][date]
                    total = list[0]+list[1]+list[2]
                    if total == 0:
                        outputSheet.cell(row=i+2, column=j+3).value = "A"
                    else:
                        outputSheet.cell(row=i+2, column=j+3).value = "P"
                        present += 1

            outputSheet.cell(row=i+2, column=last).value = len(dates)
            outputSheet.cell(row=i+2, column=last+1).value = present
            percentage_attendance = (100*present)/len(dates)
            percentage_attendance = round(percentage_attendance, 2)
            outputSheet.cell(row=i+2, column=last +
                             2).value = percentage_attendance

        outputFile.save(outputFileName)
    except:
        print("Folder named output doesn't exist")
        exit()


def roll_attendance_func():
    title = ["Date", "Roll", "Name", "Total Attendance Count",
             "Real", "Duplicate", "Invalid", "Absent"]

    for rollNum in roll_to_name.keys():
        try:
            outputFileName = "output/" + rollNum + ".xlsx"
            outputFile = openpyxl.Workbook()
            outputSheet = outputFile.active

            for i, word in enumerate(title):
                outputSheet.cell(row=1, column=i+1).value = word
            outputSheet.cell(row=2, column=2).value = rollNum
            outputSheet.cell(row=2, column=3).value = roll_to_name[rollNum]

            attendance = roll_attendance[rollNum]  # map of date -> array

            for i, date in enumerate(attendance.keys()):
                outputSheet.cell(row=3+i, column=1).value = date
                list = attendance[date]
                total = list[0]+list[1]+list[2]

                outputSheet.cell(row=3+i, column=4).value = total
                outputSheet.cell(row=3+i, column=5).value = list[0]
                outputSheet.cell(row=3+i, column=6).value = list[1]
                outputSheet.cell(row=3+i, column=7).value = list[2]

                if total == 0:
                    outputSheet.cell(row=3+i, column=8).value = 1
                else:
                    outputSheet.cell(row=3+i, column=8).value = 0

            outputFile.save(outputFileName)
        except:
            print("The folder output doesn't exist")
            exit()


def valid_day(date):
    day_rank = datetime.strptime(date, '%d-%m-%Y').weekday()

    if day_rank == 0 or day_rank == 3:
        return True
    return False


def valid_time(time):
    # e.g., time = 15:30
    hrs, minute = time.split(":")
    hrs = int(hrs)
    minute = int(minute)
    if hrs == 15 and minute == 0:
        return True
    if hrs == 14:
        return True
    return False


def attendance_count_func():
    try:
        # Opening input file
        f = open(inputAttendance, "r")

        f.readline()

        all_lines = f.readlines()

        for rolls in roll_to_name.keys():
            roll_attendance[rolls] = {}

        for line in all_lines:
            line = line.strip()
            timestamp, naming = line.split(",")
            date, time = timestamp.split(" ")
            rollNumber = naming[:8]

            if valid_day(date):
                if date not in dates:
                    dates.append(date)

                if date not in roll_attendance[rollNumber]:
                    roll_attendance[rollNumber][date] = [0, 0, 0]

                if valid_time(time):
                    if (roll_attendance[rollNumber][date][0] == 0):
                        roll_attendance[rollNumber][date][0] = 1
                    else:
                        roll_attendance[rollNumber][date][1] += 1

                else:
                    roll_attendance[rollNumber][date][2] += 1
    except FileNotFoundError:
        print('File not found')
        exit()


def map_roll_to_num_func():
    try:
        # Opening input file
        f = open(inputRegisteredFile, "r")
        # Reading label line
        f.readline()
        # Reading all lines of input
        all_lines = f.readlines()
        # Iterating all lines
        for line in all_lines:
            line = line.strip()
            list = line.split(",")

            rollNo = list[0].strip()
            name = list[1].strip()
            roll_to_name[rollNo] = name

        f.close()
    except FileNotFoundError:
        print("File not found")
        exit()


def attendance_report():
    try:
        # Method to map roll num to name
        map_roll_to_num_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()
    try:
        # Method to count attendance
        attendance_count_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()
    try:
        # Saving rollNum attendance
        roll_attendance_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()
    try:
        # Saving consolidate attendance
        consolidate_attendance_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()


try:
    attendance_report()
except NameError:
    print('Either the function is not created or the name is not correct.')
    exit()

# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))


'''


from sys import exit
import pandas as pd
import datetime
from datetime import date
import calendar
import os

dir_name=os.get_
data1 = pd.read_csv (dir_name+'/input_registered_students.csv')


data2 = pd.read_csv(dir_name+'/input_attendance.csv')

n1 = len(data1)
n2 = len(data2)
students = []
names = [""]
studentnames = {}
for i in range(n1):
    t = data1.loc[i, "Roll No"]
    s = data1.loc[i, "Name"]
    studentnames[t] = s
    names.append(s)
    students.append(t)

dates = [""]
for i in range(n2):
    t = data2.loc[i, "Timestamp"]
    t = t[:10]
    d = datetime.datetime.strptime(t, "%d-%m-%Y")
    s = d.strftime('%Y-%m-%d')
    s = datetime.datetime.strptime(s, '%Y-%m-%d')
    if(calendar.day_name[s.weekday()] == 'Monday' or calendar.day_name[s.weekday()] == 'Thursday'):
        dates.append(s)

day_delta = datetime.timedelta(days=1)
ini = dates[1]
fin = dates[len(dates)-1]+day_delta
dates = [""]

for i in range((fin - ini).days):
    s = ini + i*day_delta
    if(calendar.day_name[s.weekday()] == 'Monday' or calendar.day_name[s.weekday()] == 'Thursday'):
        s = s.strftime('%Y-%m-%d')
        d = datetime.datetime.strptime(s, "%Y-%m-%d")
        s = d.strftime('%d-%m-%Y')
        if s not in dates:
            dates.append(s)

final = pd.DataFrame()
roll = [""]
roll.extend(students)
final['Roll'] = roll
final['Name'] = names
for d in dates:
    if d != '':
        l = ['A'for i in range(n1+1)]
        l[0] = " "
        final[d] = l
l = [len(dates)-1 for i in range(n1+1)]
l[0] = " "
final['Actual Lecture Taken'] = l
l = [0 for i in range(n1+1)]
l[0] = " "
final['Total Real'] = l
l = [0 for i in range(n1+1)]
l[0] = " "
final['% Attendance'] = l
for stur in students:
    roll = [""for i in range(len(dates))]
    roll[0] = stur
    name = [""for i in range(len(dates))]
    name[0] = studentnames[stur]
    df = pd.DataFrame()
    df['Date'] = dates
    df['Roll'] = roll
    df['Name'] = name
    tac = [0 for i in range(len(dates))]
    tac[0] = ""
    real = [0 for i in range(len(dates))]
    real[0] = ""
    dup = [0 for i in range(len(dates))]
    dup[0] = ""
    invalid = [0 for i in range(len(dates))]
    invalid[0] = ""
    absent = [1 for i in range(len(dates))]
    absent[0] = ""
    j = students.index(stur)+1
    for i in range(n2):
        t = data2.loc[i, "Timestamp"]
        t = t[:10]
        r = data2.loc[i, "Attendance"]
        r = r[:8]
        if t in dates and r == stur:
            ind = dates.index(t)
            tac[ind] += 1
            time = data2.loc[i, "Timestamp"]
            time = time[11:]
            if(time[0] == '1' and time[1] == '4') or time[:5] == '15:00':
                if real[ind] == 0:
                    real[ind] = 1
                    absent[ind] = 0
                    final.loc[j, t] = 'P'
                    final.loc[j, 'Total Real'] += 1
                else:
                    dup[ind] += 1
            else:
                invalid[ind] += 1
    for i in range(1, n1+1):
        a = final.loc[i, 'Total Real']
        b = len(dates)-1
        if(a != 0):
            c = round(a/b*100, 2)
            final.loc[i, '% Attendance'] = c
    df['Total Attendance Count'] = tac
    df['Real'] = real
    df['duplicate'] = dup
    df['Invalid'] = invalid
    df['Absent'] = absent
    t = dir_name+"/output/"+stur+".xlsx"
    df.to_excel(t, index=False)
final.to_excel(dir_name+'/output/attendance_report_consolidated.xlsx', index=False)
