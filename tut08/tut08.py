import os
import pandas as pd
import openpyxl
from datetime import datetime
start_time = datetime.now()
'''
#Help
def scorecard():
	pass
###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


scorecard()

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
'''
dir_name = os.path.dirname(__file__)


start_time = datetime.now()
innings_india = open(dir_name+"/india_inns2.txt", "r+")  # india batting
innings_pakistan = open(dir_name+"/pak_inns1.txt", "r+")  # pakistan batting
teams = open(dir_name+"/teams.txt", "r+")
input_team = teams.readlines()

pak_team = input_team[0]
pak_players = pak_team[23:-1:].split(",")

ind_team = input_team[2]
ind_players = ind_team[20:-1:].split(",")


lst_ind = innings_india.readlines()  # 124
for i in lst_ind:
    if i == '\n':
        lst_ind.remove(i)


lst_pak = innings_pakistan.readlines()  # 123
for i in lst_pak:
    if i == '\n':
        lst_pak.remove(i)

wb = openpyxl.Workbook()
sheet = wb.active

# batting [runs,ball,4s,6s,sr]
# bowling [over,medan,runs,Wickets, NB, WD, ECO]
ind_fall_of_wickets = 0
pak_fall_of_wickets = 0
out_pak_bat = {}
ind_bowlers = {}
ind_bats = {}
pak_bats = {}
pak_bowlers = {}
pak_byes = 0
pak_bowlers_total = 0
for l in lst_pak:
    x = l.index(".")
    over_pak = l[0:x+2]
    temp = l[x+2::].split(",")
    curr_ball = temp[0].split("to")  # 0 2

    if f"{curr_ball[0].strip()}" not in ind_bowlers.keys():
        # [over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
        ind_bowlers[f"{curr_ball[0].strip()}"] = [1, 0, 0, 0, 0, 0, 0]
    elif "wide" in temp[1]:
        pass
    elif "bye" in temp[1]:
        if "FOUR" in temp[2]:
            pak_byes += 4
        elif "1" in temp[2]:
            pak_byes += 1
        elif "2" in temp[2]:
            pak_byes += 2
        elif "3" in temp[2]:
            pak_byes += 3
        elif "4" in temp[2]:
            pak_byes += 4
        elif "5" in temp[2]:
            pak_byes += 5

    else:
        ind_bowlers[f"{curr_ball[0].strip()}"][0] += 1

    if f"{curr_ball[1].strip()}" not in pak_bats.keys() and temp[1] != "wide":
        pak_bats[f"{curr_ball[1].strip()}"] = [0, 1, 0, 0,
                                               0]  # [runs,ball,4s,6s,sr]
    elif "wide" in temp[1]:
        pass
    else:
        pak_bats[f"{curr_ball[1].strip()}"][1] += 1

    if "out" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][3] += 1
        if "Bowled" in temp[1].split("!!")[0]:
            out_pak_bat[f"{curr_ball[1].strip()}"] = ("b" + curr_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w = (temp[1].split("!!")[0]).split("by")
            out_pak_bat[f"{curr_ball[1].strip()}"] = (
            	"c" + w[1] + " b " + curr_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            out_pak_bat[f"{curr_ball[1].strip()}"] = ("lbw  b "+curr_ball[0])

    if "no run" in temp[1] or "out" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 0
        pak_bats[f"{curr_ball[1].strip()}"][0] += 0
    elif "1 run" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 1
        pak_bats[f"{curr_ball[1].strip()}"][0] += 1
    elif "2 run" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 2
        pak_bats[f"{curr_ball[1].strip()}"][0] += 2
    elif "3 run" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 3
        pak_bats[f"{curr_ball[1].strip()}"][0] += 3
    elif "4 run" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        pak_bats[f"{curr_ball[1].strip()}"][0] += 4
    elif "FOUR" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        pak_bats[f"{curr_ball[1].strip()}"][0] += 4
        pak_bats[f"{curr_ball[1].strip()}"][2] += 1
    elif "SIX" in temp[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 6
        pak_bats[f"{curr_ball[1].strip()}"][0] += 6
        pak_bats[f"{curr_ball[1].strip()}"][3] += 1
    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            # print(temp[1][1])
            ind_bowlers[f"{curr_ball[0].strip()}"][2] += int(temp[1][1])
            ind_bowlers[f"{curr_ball[0].strip()}"][5] += int(temp[1][1])
        else:
            ind_bowlers[f"{curr_ball[0].strip()}"][2] += 1
            ind_bowlers[f"{curr_ball[0].strip()}"][5] += 1

for val in pak_bats.values():
    val[-1] = round((val[0]/val[1])*100, 2)


############# india innings ##############
ind_bowlers_total = 0
ind_byes = 0
out_ind_bat = {}
for l in lst_ind:
    x = l.index(".")
    over_ind = l[0:x+2]

    temp = l[x+2::].split(",")

    curr_ball = temp[0].split("to")  # 0 2
    if f"{curr_ball[0].strip()}" not in pak_bowlers.keys():
        # [over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
        pak_bowlers[f"{curr_ball[0].strip()}"] = [1, 0, 0, 0, 0, 0, 0]
    elif "wide" in temp[1]:
        pass
    elif "bye" in temp[1]:
        if "FOUR" in temp[2]:
            ind_byes += 4
        elif "1" in temp[2]:
            ind_byes += 1
        elif "2" in temp[2]:
            ind_byes += 2
        elif "3" in temp[2]:
            ind_byes += 3
        elif "4" in temp[2]:
            ind_byes += 4
        elif "5" in temp[2]:
            ind_byes += 5
    else:
        pak_bowlers[f"{curr_ball[0].strip()}"][0] += 1

    if f"{curr_ball[1].strip()}" not in ind_bats.keys() and temp[1] != "wide":
        ind_bats[f"{curr_ball[1].strip()}"] = [0, 1, 0, 0,
                                               0]  # [runs,ball,4s,6s,sr]
    elif "wide" in temp[1]:
        pass
    else:
        ind_bats[f"{curr_ball[1].strip()}"][1] += 1

    if "out" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][3] += 1

        if "Bowled" in temp[1].split("!!")[0]:
            out_ind_bat[f"{curr_ball[1].strip()}"] = ("b" + curr_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w = (temp[1].split("!!")[0]).split("by")
            out_ind_bat[f"{curr_ball[1].strip()}"] = (
            	"c" + w[1] + " b " + curr_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            out_ind_bat[f"{curr_ball[1].strip()}"] = ("lbw  b "+curr_ball[0])

    if "no run" in temp[1] or "out" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 0
        ind_bats[f"{curr_ball[1].strip()}"][0] += 0
    elif "1 run" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 1
        ind_bats[f"{curr_ball[1].strip()}"][0] += 1
    elif "2 run" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 2
        ind_bats[f"{curr_ball[1].strip()}"][0] += 2
    elif "3 run" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 3
        ind_bats[f"{curr_ball[1].strip()}"][0] += 3
    elif "4 run" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        ind_bats[f"{curr_ball[1].strip()}"][0] += 4
    elif "FOUR" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        ind_bats[f"{curr_ball[1].strip()}"][0] += 4
        ind_bats[f"{curr_ball[1].strip()}"][2] += 1
    elif "SIX" in temp[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 6
        ind_bats[f"{curr_ball[1].strip()}"][0] += 6
        ind_bats[f"{curr_ball[1].strip()}"][3] += 1
    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            pak_bowlers[f"{curr_ball[0].strip()}"][2] += int(temp[1][1])
            pak_bowlers[f"{curr_ball[0].strip()}"][5] += int(temp[1][1])
        else:
            pak_bowlers[f"{curr_ball[0].strip()}"][2] += 1
            pak_bowlers[f"{curr_ball[0].strip()}"][5] += 1


for val in ind_bats.values():
    val[-1] = round((val[0]/val[1])*100, 2)

for val in pak_bats.values():
    val[-1] = round((val[0]/val[1])*100, 2)

for val in ind_bowlers.values():
    if val[0] % 6 == 0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0] % 6)/10

for val in pak_bowlers.values():
    if val[0] % 6 == 0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0] % 6)/10

for val in ind_bowlers.values():  # economy
    x = str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1] = round((val[2]/balls)*6, 1)
    else:
        val[-1] = round((val[2]/val[0]), 1)


for val in pak_bowlers.values():  # economy
    x = str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1] = round((val[2]/balls)*6, 1)
    else:
        val[-1] = round((val[2]/val[0]), 1)


# pakistan batting
bat_pak_name = []
for key in pak_bats.keys():
    bat_pak_name.append(key)


for i in range(len(pak_bats)):
    sheet.cell(5+i, 1).value = bat_pak_name[i]
    sheet.cell(5+i, 5).value = pak_bats[bat_pak_name[i]][0]
    sheet.cell(5+i, 6).value = pak_bats[bat_pak_name[i]][1]
    sheet.cell(5+i, 7).value = pak_bats[bat_pak_name[i]][2]
    sheet.cell(5+i, 8).value = pak_bats[bat_pak_name[i]][3]
    sheet.cell(5+i, 9).value = pak_bats[bat_pak_name[i]][4]
    if bat_pak_name[i] not in out_pak_bat:
        sheet.cell(5+i, 3).value = "not out"
    else:
        sheet.cell(5+i, 3).value = out_pak_bat[bat_pak_name[i]]

sheet.cell(3, 1).value = "BATTERS"
sheet["E3"] = "RUNS"
sheet["F3"] = "BALLS"
sheet["G3"] = " 4s "
sheet["H3"] = " 6s "
sheet["I3"] = "  SR  "

# india bowling

sheet["A18"] = "BOWLER"
sheet["C18"] = "OVER"
sheet["D18"] = "MAIDEN"
sheet["E18"] = "RUNS"
sheet["F18"] = "WICKET"
sheet["G18"] = "NO-BALL"
sheet["H18"] = "WIDE"
sheet["I18"] = "ECONOMY"

names_pak_bowlers = []
for key in pak_bowlers.keys():
    names_pak_bowlers.append(key)

for i in range(len(pak_bowlers)):
    sheet.cell(42+i, 1).value = names_pak_bowlers[i]
    sheet.cell(42+i, 3).value = pak_bowlers[names_pak_bowlers[i]][0]
    sheet.cell(42+i, 4).value = pak_bowlers[names_pak_bowlers[i]][1]
    sheet.cell(42+i, 5).value = pak_bowlers[names_pak_bowlers[i]][2]
    sheet.cell(42+i, 6).value = pak_bowlers[names_pak_bowlers[i]][3]
    sheet.cell(42+i, 7).value = pak_bowlers[names_pak_bowlers[i]][4]
    sheet.cell(42+i, 8).value = pak_bowlers[names_pak_bowlers[i]][5]
    sheet.cell(42+i, 9).value = pak_bowlers[names_pak_bowlers[i]][6]
    pak_bowlers_total += pak_bowlers[names_pak_bowlers[i]][2]
    ind_fall_of_wickets += pak_bowlers[names_pak_bowlers[i]][3]

# india batting
sheet.cell(11+len(pak_bats)+len(pak_bowlers), 1).value = "# INDIA"
sheet.cell(11+len(pak_bats)+len(pak_bowlers), 2).value = " INNINGS"

names_ind_batters = []
for key in ind_bats.keys():
    names_ind_batters.append(key)


for i in range(len(ind_bats)):
    sheet.cell(31+i, 1).value = names_ind_batters[i]
    sheet.cell(31+i, 5).value = ind_bats[names_ind_batters[i]][0]
    sheet.cell(31+i, 6).value = ind_bats[names_ind_batters[i]][1]
    sheet.cell(31+i, 7).value = ind_bats[names_ind_batters[i]][2]
    sheet.cell(31+i, 8).value = ind_bats[names_ind_batters[i]][3]
    sheet.cell(31+i, 9).value = ind_bats[names_ind_batters[i]][4]

    if names_ind_batters[i] not in out_ind_bat:
        sheet.cell(31+i, 3).value = "not out"
    else:
        sheet.cell(31+i, 3).value = out_ind_bat[names_ind_batters[i]]

sheet["A29"] = "BATTERS"
sheet["E29"] = "RUNS"
sheet["F29"] = "BALLS"
sheet["G29"] = " 4s "
sheet["H29"] = " 6s "
sheet["I29"] = "  SR  "

# india bowling

sheet["A40"] = "BOWLER"
sheet["C40"] = "OVER"
sheet["D40"] = "MAIDEN"
sheet["E40"] = "RUNS"
sheet["F40"] = "WICKET"
sheet["G40"] = "NO-BALL"
sheet["H40"] = "WIDE"
sheet["I40"] = "ECONOMY"

names_ind_bowlers = []
for key in ind_bowlers.keys():
    names_ind_bowlers.append(key)

for i in range(len(ind_bowlers)):

    sheet.cell(20+i, 1).value = names_ind_bowlers[i]
    sheet.cell(20+i, 3).value = ind_bowlers[names_ind_bowlers[i]][0]
    sheet.cell(20+i, 4).value = ind_bowlers[names_ind_bowlers[i]][1]
    sheet.cell(20+i, 5).value = ind_bowlers[names_ind_bowlers[i]][2]
    sheet.cell(20+i, 6).value = ind_bowlers[names_ind_bowlers[i]][3]
    sheet.cell(20+i, 7).value = ind_bowlers[names_ind_bowlers[i]][4]
    sheet.cell(20+i, 8).value = ind_bowlers[names_ind_bowlers[i]][5]
    sheet.cell(20+i, 9).value = ind_bowlers[names_ind_bowlers[i]][6]
    ind_bowlers_total += ind_bowlers[names_ind_bowlers[i]][2]
    pak_fall_of_wickets += ind_bowlers[names_ind_bowlers[i]][3]

ind_total_score = ind_bowlers_total+pak_byes
pak_total_score = pak_bowlers_total+ind_byes
# print(over_ind)
# print(over_pak)
sheet["E27"] = " "+str(ind_total_score) + " - " + str(ind_fall_of_wickets)
sheet["F27"] = str(over_ind)
Eone = " "+str(pak_total_score) + " - " + str(pak_fall_of_wickets)
Fone = str(over_pak)

wb.save(dir_name+"/Scoreboard.xlsx")

df = pd.read_excel(dir_name+'/Scoreboard.xlsx')

df = df.set_axis(['PAKISTAN', ' INNINGS'] + [" ", " ",
                 Eone, Fone, " ", " ", " "], axis='columns')

df.to_csv(dir_name+'/Scorecard.csv', index=False)

try:
    os.path.exists(dir_name+"/Scoreboard.xlsx")
    os.remove(dir_name+"/Scoreboard.xlsx")  # deleting output excel
except:
    print("Extra created file does not exist")

end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
